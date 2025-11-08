from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from io import BytesIO
import logging

logger = logging.getLogger(__name__)

def gerar_xlsx(processos: list, execucao: dict) -> BytesIO:
    """Gera arquivo XLSX com os processos de indeferimento"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Semana {execucao['semana']}"
    
    # Cabeçalho
    headers = ['EMAIL', 'MARCA', 'PROCESSO']
    ws.append(headers)
    
    # Estilizar cabeçalho (cor laranja InHands)
    header_fill = PatternFill(start_color='FE7C1F', end_color='FE7C1F', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Dados
    for processo in processos:
        ws.append([
            processo.get('email', ''),
            processo.get('marca', ''),
            processo.get('numero_processo', '')
        ])
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 25
    
    # Salvar em buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    logger.info(f"XLSX gerado com {len(processos)} processos")
    return buffer